#!/usr/bin/env python3
"""
Script d'extraction automatique des données V1-V5 depuis les fichiers Excel Monka
Génère des fichiers JSON lisibles pour chaque vulnérabilité trouvée
"""

import json
import os
from openpyxl import load_workbook

# Chemins des fichiers Excel
EXCEL_DIR = "/Users/antonin/monka/SOURCES/excel"
OUTPUT_DIR = "/Users/antonin/monka/SOURCES/extracted"

# Fichiers Excel à scanner
EXCEL_FILES = [
    "Tableau SOPHIE CAT + Reco-Nouveau ques tionnaire par Vulnérabilité(1) (7).xlsx",
    "Questionnaire_Etienne_1258.xlsx",
    "microparcours_aidant.xlsx",
    "Questionnaire_Etienne_1258-1_suivi_mensuel.xlsx"
]

def ensure_output_dir():
    """Crée le dossier de sortie s'il n'existe pas"""
    os.makedirs(OUTPUT_DIR, exist_ok=True)

def scan_excel_file(filepath):
    """Scanne un fichier Excel et retourne la structure complète"""
    print(f"\n{'='*60}")
    print(f"📊 Analyse de: {os.path.basename(filepath)}")
    print(f"{'='*60}")
    
    try:
        wb = load_workbook(filepath, data_only=True)
        
        file_data = {
            "filename": os.path.basename(filepath),
            "sheets": {}
        }
        
        for sheet_name in wb.sheetnames:
            print(f"\n📑 Feuille: {sheet_name}")
            ws = wb[sheet_name]
            
            # Récupérer les données de la feuille
            sheet_data = {
                "name": sheet_name,
                "rows": ws.max_row,
                "cols": ws.max_column,
                "headers": [],
                "data": []
            }
            
            # Lire les headers (première ligne)
            headers = []
            for col in range(1, min(ws.max_column + 1, 20)):  # Max 20 colonnes
                cell = ws.cell(row=1, column=col)
                headers.append(str(cell.value) if cell.value else f"Col_{col}")
            sheet_data["headers"] = headers
            print(f"   Headers: {headers[:5]}{'...' if len(headers) > 5 else ''}")
            
            # Lire les données (lignes 2+)
            for row in range(2, min(ws.max_row + 1, 200)):  # Max 200 lignes
                row_data = {}
                for col in range(1, min(ws.max_column + 1, 20)):
                    cell = ws.cell(row=row, column=col)
                    col_name = headers[col-1] if col-1 < len(headers) else f"Col_{col}"
                    row_data[col_name] = str(cell.value) if cell.value else None
                
                # Ignorer les lignes vides
                if any(v for v in row_data.values()):
                    sheet_data["data"].append(row_data)
            
            print(f"   Lignes de données: {len(sheet_data['data'])}")
            
            file_data["sheets"][sheet_name] = sheet_data
        
        wb.close()
        return file_data
        
    except Exception as e:
        print(f"❌ Erreur: {str(e)}")
        return None

def identify_vulnerability_sheets(file_data):
    """Identifie les feuilles correspondant à chaque vulnérabilité"""
    vulnerability_mapping = {
        "V1": ["social", "relationnel", "v1", "soutien", "entourage"],
        "V2": ["fragilité", "proche", "v2", "vulnérabilité"],
        "V3": ["santé", "aidant", "v3", "epuisement", "fatigue"],
        "V4": ["médical", "parcours", "v4", "maladie", "pathologie"],
        "V5": ["administratif", "v5", "démarche", "droit"]
    }
    
    found = {}
    
    for sheet_name, sheet_data in file_data.get("sheets", {}).items():
        sheet_lower = sheet_name.lower()
        
        for vuln, keywords in vulnerability_mapping.items():
            if any(kw in sheet_lower for kw in keywords):
                if vuln not in found:
                    found[vuln] = []
                found[vuln].append({
                    "sheet": sheet_name,
                    "rows": len(sheet_data.get("data", [])),
                    "headers": sheet_data.get("headers", [])
                })
    
    return found

def extract_questions_from_sheet(sheet_data):
    """Extrait les questions d'une feuille de données"""
    questions = []
    headers = sheet_data.get("headers", [])
    
    # Chercher les colonnes pertinentes
    id_col = None
    text_col = None
    score_col = None
    reco_col = None
    
    for i, h in enumerate(headers):
        h_lower = str(h).lower() if h else ""
        if any(x in h_lower for x in ["id", "code", "num"]):
            id_col = i
        if any(x in h_lower for x in ["question", "libellé", "texte", "intitulé"]):
            text_col = i
        if any(x in h_lower for x in ["score", "point", "valeur"]):
            score_col = i
        if any(x in h_lower for x in ["reco", "recommandation", "conseil"]):
            reco_col = i
    
    print(f"   Colonnes détectées: ID={id_col}, Text={text_col}, Score={score_col}, Reco={reco_col}")
    
    for row in sheet_data.get("data", []):
        values = list(row.values())
        
        question = {
            "id": values[id_col] if id_col is not None and id_col < len(values) else None,
            "text": values[text_col] if text_col is not None and text_col < len(values) else None,
            "score": values[score_col] if score_col is not None and score_col < len(values) else None,
            "recommendation": values[reco_col] if reco_col is not None and reco_col < len(values) else None,
            "raw_data": row
        }
        
        if question["text"] or question["id"]:
            questions.append(question)
    
    return questions

def main():
    print("🚀 Démarrage de l'extraction Excel Monka")
    print(f"📁 Dossier source: {EXCEL_DIR}")
    print(f"📁 Dossier sortie: {OUTPUT_DIR}")
    
    ensure_output_dir()
    
    all_data = {}
    vulnerability_summary = {}
    
    for excel_file in EXCEL_FILES:
        filepath = os.path.join(EXCEL_DIR, excel_file)
        
        if not os.path.exists(filepath):
            print(f"\n⚠️  Fichier non trouvé: {excel_file}")
            continue
        
        file_data = scan_excel_file(filepath)
        
        if file_data:
            all_data[excel_file] = file_data
            
            # Identifier les vulnérabilités
            vulns = identify_vulnerability_sheets(file_data)
            if vulns:
                print(f"\n🎯 Vulnérabilités trouvées:")
                for v, sheets in vulns.items():
                    print(f"   {v}: {[s['sheet'] for s in sheets]}")
                    if v not in vulnerability_summary:
                        vulnerability_summary[v] = []
                    vulnerability_summary[v].extend(sheets)
            
            # Sauvegarder les données brutes du fichier
            output_file = os.path.join(OUTPUT_DIR, f"{os.path.splitext(excel_file)[0]}_raw.json")
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(file_data, f, ensure_ascii=False, indent=2)
            print(f"\n💾 Sauvegardé: {output_file}")
    
    # Résumé final
    print(f"\n{'='*60}")
    print("📊 RÉSUMÉ DE L'EXTRACTION")
    print(f"{'='*60}")
    
    print(f"\n📁 Fichiers analysés: {len(all_data)}")
    
    print(f"\n🎯 Vulnérabilités identifiées:")
    for v in ["V1", "V2", "V3", "V4", "V5"]:
        if v in vulnerability_summary:
            sheets = vulnerability_summary[v]
            total_rows = sum(s['rows'] for s in sheets)
            print(f"   {v}: {len(sheets)} feuille(s), {total_rows} lignes de données")
        else:
            print(f"   {v}: ❌ Non trouvée")
    
    # Sauvegarder le résumé
    summary_file = os.path.join(OUTPUT_DIR, "extraction_summary.json")
    with open(summary_file, 'w', encoding='utf-8') as f:
        json.dump({
            "files_analyzed": list(all_data.keys()),
            "vulnerabilities": vulnerability_summary,
            "raw_files": [f"{os.path.splitext(f)[0]}_raw.json" for f in all_data.keys()]
        }, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Extraction terminée!")
    print(f"📄 Résumé: {summary_file}")
    print(f"📁 Données brutes: {OUTPUT_DIR}/")

if __name__ == "__main__":
    main()
