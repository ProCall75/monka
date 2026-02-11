"""
Met à jour le fichier Excel Questionnaire_Monka_Complet.xlsx :
1. Supprime la colonne "Aidance" (si encore présente)
2. Ajoute / met à jour la colonne "Typage" (scorante / déclenchante / scorante + déclenchante / informative / trigger)

Sources:
- Scorantes : fichiers scoring.md legacy (V1-V5) = 37 questions
- Déclenchantes : Supabase activation_rules (69 questions uniques, incluant critique + CCC + standard)
"""

import openpyxl

# --- 1. Questions scorantes (extraites des fichiers scoring.md legacy) ---
SCORANTES = {
    # V1 (8 questions, score max 15)
    "E1", "E2", "E4", "E5", "N20", "O27", "O28", "O30",
    # V2 (18 questions)
    "E18", "E25", "E26", "E36", "E37", "E43", "N11", "N12", "N13", "N24", "N34",
    "O4", "O6", "O7", "O13", "O26", "O29", "O44",
    # V3 (6 questions)
    "E7", "E8", "E9", "E10", "E11", "O33",
    # V4 (3 questions)
    "E47", "E54", "E57",
    # V5 (3 questions)
    "E66", "E69", "E70",
}  # Total: 38

# --- 2. Questions déclenchantes — TOUTES les question_ids dans activation_rules ---
# Source: SELECT DISTINCT unnest(question_ids) FROM activation_rules (69 résultats)
DECLENCHANTES = {
    "E1", "E11", "E12", "E13", "E14", "E15", "E16", "E18",
    "E2", "E21", "E23", "E24", "E25", "E26", "E27", "E28",
    "E36", "E37", "E38", "E4", "E40", "E42", "E43", "E44",
    "E45", "E46", "E47", "E5", "E50", "E51", "E52", "E54",
    "E57", "E6", "E61", "E62", "E68", "E7", "E8", "E9",
    "N11", "N12", "N13", "N20", "N21", "N22", "N25", "N34",
    "N38", "N39", "N4", "N7", "N8", "N9",
    "O13", "O24", "O27", "O28", "O30", "O31", "O32",
    "O4", "O44", "O48", "O49", "O51", "O53", "O8", "O9",
}  # Total: 69

# --- 3. Triggers ---
TRIGGERS = {"N3", "O35", "O36", "N1", "O64", "O46", "O14", "O1", "O63", "N26", "E71", "E72", "O2", "N31", "O49"}

def get_typing(qid):
    if qid in TRIGGERS:
        return "trigger"
    is_scorante = qid in SCORANTES
    is_declenchante = qid in DECLENCHANTES
    if is_scorante and is_declenchante:
        return "scorante + déclenchante"
    elif is_scorante:
        return "scorante"
    elif is_declenchante:
        return "déclenchante"
    else:
        return "informative"

# --- 4. Modify Excel ---
INPUT = "/Users/antonin/monka/LIVRABLES/Questionnaire_Monka_Complet.xlsx"

wb = openpyxl.load_workbook(INPUT)
ws = wb[wb.sheetnames[0]]

headers = [cell.value for cell in ws[1]]
print(f"Colonnes actuelles: {headers}")

# Remove Aidance if still present
if "Aidance" in headers:
    aidance_col = headers.index("Aidance") + 1
    ws.delete_cols(aidance_col)
    print(f"Colonne 'Aidance' supprimée")
    headers = [cell.value for cell in ws[1]]

# If Typage already exists, find it; otherwise insert after Classification
if "Typage" in headers:
    typage_col = headers.index("Typage") + 1
    print(f"Colonne 'Typage' existe déjà (col {typage_col}), mise à jour")
else:
    classification_col = headers.index("Classification") + 1
    typage_col = classification_col + 1
    ws.insert_cols(typage_col)
    ws.cell(row=1, column=typage_col, value="Typage")
    print(f"Colonne 'Typage' ajoutée (col {typage_col})")
    headers = [cell.value for cell in ws[1]]

# Fill typing
id_col = headers.index("ID") + 1
stats = {}
for row in range(2, ws.max_row + 1):
    qid = ws.cell(row=row, column=id_col).value
    if qid:
        typing = get_typing(qid)
        ws.cell(row=row, column=typage_col, value=typing)
        stats[typing] = stats.get(typing, 0) + 1

# Also update Triggers sheet
if "Triggers (15)" in wb.sheetnames:
    ws_t = wb["Triggers (15)"]
    th = [cell.value for cell in ws_t[1]]
    if "Aidance" in th:
        ws_t.delete_cols(th.index("Aidance") + 1)
        th = [cell.value for cell in ws_t[1]]
    if "Typage" not in th and "Classification" in th:
        tc = th.index("Classification") + 1
        ws_t.insert_cols(tc + 1)
        ws_t.cell(row=1, column=tc + 1, value="Typage")
        for row in range(2, ws_t.max_row + 1):
            ws_t.cell(row=row, column=tc + 1, value="trigger")

wb.save(INPUT)
print(f"\n✅ Excel sauvegardé: {INPUT}")
print(f"\nStatistiques typage (150 questions):")
for k, v in sorted(stats.items()):
    print(f"  {k}: {v}")
print(f"  Total: {sum(stats.values())}")
